"""
main.py - FastAPI application entrypoint
"""
import os
import logging
import time
from contextlib import asynccontextmanager
from typing import Optional

from dotenv import load_dotenv
load_dotenv()

from alembic.config import Config as AlembicConfig
from alembic import command as alembic_command
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, Body, Request
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import get_db
from models.inventory import Base, InventoryItem, Category, Vendor, Receipt, VendorType, ItemHistory
from models.user import User
from services.inventory_service import InventoryService
from utils.logging_config import setup_logging
from vendors.home_depot import HomeDepotAdapter
from vendors.lowes import LowesAdapter
from vendors.amazon import AmazonAdapter
from utils.ocr_service import OCRService, OCRProvider
from utils.barcode_service import lookup_barcode, decode_barcode_image, decode_barcode_via_vision
from dependencies import get_current_user
from routers.auth import router as auth_router

# ─────────────────────────────────────────────────────────────────────────────
setup_logging()
logger = logging.getLogger(__name__)

HD_API_KEY     = os.getenv("HOME_DEPOT_API_KEY", "")
LOWES_API_KEY  = os.getenv("LOWES_API_KEY", "")
AMAZON_API_KEY = os.getenv("AMAZON_API_KEY", "")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
GROQ_API_KEY   = os.getenv("GROQ_API_KEY", "")


@asynccontextmanager
async def lifespan(app: FastAPI):
    alembic_cfg = AlembicConfig("alembic.ini")
    alembic_command.upgrade(alembic_cfg, "head")
    yield


app = FastAPI(title="Inventory Manager API", version="1.0.0", lifespan=lifespan)

# CORS_ORIGINS env var: comma-separated list of allowed origins.
# Example: "https://myapp.up.railway.app,https://myapp.vercel.app"
# Falls back to localhost origins for local development.
_cors_env = os.getenv("CORS_ORIGINS", "")
CORS_ORIGINS = (
    [o.strip() for o in _cors_env.split(",") if o.strip()]
    or ["http://localhost:5173", "http://localhost:3000",
        "capacitor://localhost", "http://localhost"]   # last two for mobile later
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(auth_router, prefix="/api")


@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = time.perf_counter()
    response = await call_next(request)
    ms = (time.perf_counter() - start) * 1000
    logger.info("%s %s %d  %.0fms", request.method, request.url.path, response.status_code, ms)
    return response


try:
    app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")
except Exception:
    pass


# ── Vendor helpers ────────────────────────────────────────────────────────────

def _all_adapters() -> list:
    adapters = []
    if HD_API_KEY:
        adapters.append(HomeDepotAdapter(api_key=HD_API_KEY))
    if LOWES_API_KEY:
        adapters.append(LowesAdapter(api_key=LOWES_API_KEY))
    if AMAZON_API_KEY:
        adapters.append(AmazonAdapter(api_key=AMAZON_API_KEY))
    return adapters


def _adapter_for(vendor: str):
    if vendor == "lowes":
        if not LOWES_API_KEY:
            raise HTTPException(503, "LOWES_API_KEY not configured")
        return LowesAdapter(api_key=LOWES_API_KEY)
    if vendor == "hd":
        if not HD_API_KEY:
            raise HTTPException(503, "HOME_DEPOT_API_KEY not configured")
        return HomeDepotAdapter(api_key=HD_API_KEY)
    if vendor == "amazon":
        if not AMAZON_API_KEY:
            raise HTTPException(503, "AMAZON_API_KEY not configured")
        return AmazonAdapter(api_key=AMAZON_API_KEY)
    raise HTTPException(400, f"Unknown vendor '{vendor}' — use 'hd', 'lowes', or 'amazon'")


def _build_ocr() -> Optional[OCRService]:
    if GROQ_API_KEY:
        return OCRService(provider=OCRProvider.GROQ, api_key=GROQ_API_KEY)
    if GEMINI_API_KEY:
        return OCRService(provider=OCRProvider.GEMINI, api_key=GEMINI_API_KEY)
    if OPENAI_API_KEY:
        return OCRService(provider=OCRProvider.OPENAI, api_key=OPENAI_API_KEY)
    return None


def get_service(
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
) -> InventoryService:
    first_adapter = _all_adapters()[0] if _all_adapters() else None
    return InventoryService(db=db, vendor_adapter=first_adapter, ocr_service=_build_ocr(), user_id=_user.id, user_email=_user.email)


# ── Pydantic schemas ─────────────────────────────────────────────────────────

class ItemCreate(BaseModel):
    name:                str
    sku:                 Optional[str] = None
    description:         Optional[str] = None
    quantity:            int           = 0
    unit_price:          Optional[float] = None
    category_id:         Optional[str] = None
    vendor_id:           Optional[str] = None
    low_stock_threshold: int           = 5
    image_url:           Optional[str] = None

class ItemUpdate(ItemCreate):
    name: Optional[str] = None

class QuantityAdjust(BaseModel):
    delta: int

class CategoryCreate(BaseModel):
    name:        str
    description: Optional[str] = None
    parent_id:   Optional[str] = None

class ReceiptConfirm(BaseModel):
    line_item_ids: list[str]

class BulkDelete(BaseModel):
    ids: list[str]

class BulkQuantityAdjust(BaseModel):
    ids:   list[str]
    delta: int

class BulkCategoryAssign(BaseModel):
    ids:         list[str]
    category_id: Optional[str] = None

class CategoryUpdate(BaseModel):
    name:        Optional[str] = None
    description: Optional[str] = None
    parent_id:   Optional[str] = None

DEFAULT_CATEGORIES = [
    "Power Tools", "Hand Tools", "Fasteners", "Plumbing", "Electrical",
    "Daily Use", "Hardware", "Safety", "Paint", "Storage",
]


# ── Items ─────────────────────────────────────────────────────────────────────

@app.get("/api/items")
def list_items(
    search:         Optional[str] = Query(None),
    category_id:    Optional[str] = Query(None),
    low_stock_only: bool          = Query(False),
    skip:           int           = Query(0, ge=0),
    limit:          int           = Query(100, le=500),
    svc: InventoryService = Depends(get_service),
):
    return svc.get_all_items(
        search=search, category_id=category_id,
        low_stock_only=low_stock_only, skip=skip, limit=limit,
    )


@app.post("/api/items", status_code=201)
def create_item(payload: ItemCreate, svc: InventoryService = Depends(get_service)):
    return svc.create_item(payload.model_dump(exclude_none=True)).to_dict()


@app.get("/api/items/export")
def export_items(
    fmt:            str           = Query("csv", alias="format", pattern="^(csv|xlsx)$"),
    search:         Optional[str] = Query(None),
    category_id:    Optional[str] = Query(None),
    low_stock_only: bool          = Query(False),
    svc:            InventoryService = Depends(get_service),
):
    items = svc.get_all_items(
        search=search, category_id=category_id,
        low_stock_only=low_stock_only, skip=0, limit=10_000,
    ).get("items", [])

    headers = ["Name", "SKU", "Category", "Quantity", "Unit Price",
               "Low Stock Threshold", "Status", "Description", "Created At"]

    def row(item):
        return [
            item.get("name", ""),
            item.get("sku") or "",
            item.get("category", {}).get("name", "") if item.get("category") else "",
            item.get("quantity", 0),
            item.get("unit_price") or "",
            item.get("low_stock_threshold", 5),
            "Low Stock" if item.get("is_low_stock") else "OK",
            item.get("description") or "",
            item.get("created_at", ""),
        ]

    if fmt == "csv":
        import csv, io
        buf = io.StringIO()
        w   = csv.writer(buf)
        w.writerow(headers)
        for item in items:
            w.writerow(row(item))
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=inventory.csv"},
        )

    # xlsx
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Inventory"

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, item in enumerate(items, 2):
        for col, val in enumerate(row(item), 1):
            ws.cell(row=r, column=col, value=val)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col
        ) + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory.xlsx"},
    )


@app.get("/api/items/{item_id}")
def get_item(item_id: str, svc: InventoryService = Depends(get_service)):
    item = svc.get_item(item_id)
    if not item:
        raise HTTPException(404, "Item not found")
    return item.to_dict()


@app.patch("/api/items/{item_id}")
def update_item(item_id: str, payload: ItemUpdate, svc: InventoryService = Depends(get_service)):
    item = svc.update_item(item_id, payload.model_dump(exclude_none=True))
    if not item:
        raise HTTPException(404, "Item not found")
    return item.to_dict()


@app.delete("/api/items/{item_id}", status_code=204)
def delete_item(item_id: str, svc: InventoryService = Depends(get_service)):
    if not svc.delete_item(item_id):
        raise HTTPException(404, "Item not found")


@app.post("/api/items/{item_id}/adjust-quantity")
def adjust_quantity(item_id: str, payload: QuantityAdjust, svc: InventoryService = Depends(get_service)):
    item = svc.adjust_quantity(item_id, payload.delta)
    if not item:
        raise HTTPException(404, "Item not found")
    return item.to_dict()


# ── Audit history ────────────────────────────────────────────────────────────

@app.get("/api/history")
def get_global_history(
    skip:   int           = Query(0, ge=0),
    limit:  int           = Query(50, le=200),
    action: Optional[str] = Query(None),
    svc:    InventoryService = Depends(get_service),
):
    q = svc.db.query(ItemHistory).order_by(ItemHistory.created_at.desc())
    if action:
        q = q.filter(ItemHistory.action == action)
    total = q.count()
    return {"total": total, "entries": [r.to_dict() for r in q.offset(skip).limit(limit).all()]}


@app.get("/api/items/{item_id}/history")
def get_item_history(item_id: str, svc: InventoryService = Depends(get_service)):
    item = svc.get_item(item_id)
    if not item:
        raise HTTPException(404, "Item not found")
    rows = (
        svc.db.query(ItemHistory)
        .filter(ItemHistory.item_id == item_id)
        .order_by(ItemHistory.created_at.desc())
        .limit(100)
        .all()
    )
    return [r.to_dict() for r in rows]


# ── Bulk operations ───────────────────────────────────────────────────────────

@app.post("/api/items/bulk-delete")
def bulk_delete_items(payload: BulkDelete, svc: InventoryService = Depends(get_service)):
    deleted = sum(1 for id in payload.ids if svc.delete_item(id))
    return {"deleted": deleted}


@app.post("/api/items/bulk-adjust")
def bulk_adjust_quantity(payload: BulkQuantityAdjust, svc: InventoryService = Depends(get_service)):
    updated = [
        item.to_dict()
        for id in payload.ids
        if (item := svc.adjust_quantity(id, payload.delta))
    ]
    return {"updated": len(updated), "items": updated}


@app.post("/api/items/bulk-category")
def bulk_assign_category(payload: BulkCategoryAssign, svc: InventoryService = Depends(get_service)):
    updated = 0
    for id in payload.ids:
        item = svc.update_item(id, {"category_id": payload.category_id})
        if item:
            updated += 1
    return {"updated": updated}


# ── Vendor SKU lookup ──────────────────────────────────────────────────────────

@app.get("/api/vendor/sku/{sku}")
async def vendor_sku_lookup(
    sku:    str,
    vendor: Optional[str] = Query(None),
    db:     Session = Depends(get_db),
    _user:  User = Depends(get_current_user),
):
    ocr_service = _build_ocr()

    if vendor:
        adapter = _adapter_for(vendor)
        svc = InventoryService(db=db, vendor_adapter=adapter, ocr_service=ocr_service)
        try:
            data = await svc.enrich_from_vendor(sku)
        except RuntimeError as e:
            raise HTTPException(503, str(e))
        if not data:
            raise HTTPException(404, f"SKU {sku} not found at {vendor}")
        return data

    adapters = _all_adapters()
    logger.info("SKU lookup %s — %d adapter(s) available", sku, len(adapters))
    if not adapters:
        raise HTTPException(503, "No vendor API keys configured")

    for adapter in adapters:
        svc = InventoryService(db=db, vendor_adapter=adapter, ocr_service=ocr_service)
        try:
            data = await svc.enrich_from_vendor(sku)
            logger.info("Adapter %s returned: %s", adapter.__class__.__name__, data)
            if data:
                return data
        except Exception as e:
            logger.exception("Adapter %s failed for SKU %s: %s", adapter.__class__.__name__, sku, e)
            continue

    raise HTTPException(404, f"SKU {sku} not found in any vendor catalog")


# ── Barcode lookup ────────────────────────────────────────────────────────────

@app.post("/api/barcode/decode-image")
async def barcode_decode_image(
    file:  UploadFile = File(...),
    _user: User = Depends(get_current_user),
):
    contents = await file.read()
    code = decode_barcode_image(contents)
    if not code and OPENAI_API_KEY:
        code = await decode_barcode_via_vision(contents, OPENAI_API_KEY)
    if not code:
        raise HTTPException(404, "No barcode found in image")
    return {"barcode": code}


@app.get("/api/barcode/{barcode}")
async def barcode_lookup(barcode: str, _user: User = Depends(get_current_user)):
    result = await lookup_barcode(barcode, openai_api_key=OPENAI_API_KEY)
    if not result:
        raise HTTPException(404, f"No product found for barcode {barcode}")
    return result


# ── Receipt processing ─────────────────────────────────────────────────────────

@app.post("/api/receipts/upload")
async def upload_receipt(
    file:      UploadFile = File(...),
    vendor_id: Optional[str] = None,
    svc:       InventoryService = Depends(get_service),
):
    contents = await file.read()
    path     = svc.save_receipt_image(contents, file.filename or "receipt.jpg")
    try:
        receipt = await svc.process_receipt(path, vendor_id=vendor_id)
    except RuntimeError as e:
        raise HTTPException(503, f"OCR unavailable: {e}")
    if receipt.status == "failed":
        detail = receipt.raw_ocr_text or "OCR processing failed"
        raise HTTPException(422, detail)
    return receipt.to_dict()


@app.get("/api/receipts/{receipt_id}")
def get_receipt(receipt_id: str, svc: InventoryService = Depends(get_service)):
    receipt = svc.db.get(Receipt, receipt_id)
    if not receipt:
        raise HTTPException(404, "Receipt not found")
    return receipt.to_dict()


@app.post("/api/receipts/{receipt_id}/confirm")
def confirm_receipt_items(
    receipt_id: str,
    payload:    ReceiptConfirm,
    svc:        InventoryService = Depends(get_service),
):
    items = svc.create_items_from_receipt(receipt_id, payload.line_item_ids)
    return {"created": len(items), "items": [i.to_dict() for i in items]}


# ── Categories ────────────────────────────────────────────────────────────────

@app.get("/api/categories")
def list_categories(svc: InventoryService = Depends(get_service)):
    cats = svc.db.query(Category).order_by(Category.name).all()
    return [c.to_dict() for c in cats]


@app.post("/api/categories", status_code=201)
def create_category(payload: CategoryCreate, svc: InventoryService = Depends(get_service)):
    cat = Category(**payload.model_dump(exclude_none=True))
    svc.db.add(cat)
    svc.db.commit()
    svc.db.refresh(cat)
    return cat.to_dict()


@app.post("/api/categories/seed", status_code=201)
def seed_categories(svc: InventoryService = Depends(get_service)):
    existing = {c.name for c in svc.db.query(Category).all()}
    created = []
    for name in DEFAULT_CATEGORIES:
        if name not in existing:
            svc.db.add(Category(name=name))
            created.append(name)
    svc.db.commit()
    return {"created": len(created), "names": created}


@app.patch("/api/categories/{cat_id}")
def update_category(cat_id: str, payload: CategoryUpdate, svc: InventoryService = Depends(get_service)):
    cat = svc.db.get(Category, cat_id)
    if not cat:
        raise HTTPException(404, "Category not found")
    for k, v in payload.model_dump(exclude_none=True).items():
        setattr(cat, k, v)
    svc.db.commit()
    svc.db.refresh(cat)
    return cat.to_dict()


@app.delete("/api/categories/{cat_id}", status_code=204)
def delete_category(cat_id: str, svc: InventoryService = Depends(get_service)):
    cat = svc.db.get(Category, cat_id)
    if not cat:
        raise HTTPException(404, "Category not found")
    for item in cat.items:
        item.category_id = None
    svc.db.delete(cat)
    svc.db.commit()


# ── Dashboard summary ─────────────────────────────────────────────────────────

@app.get("/api/dashboard/summary")
def dashboard_summary(svc: InventoryService = Depends(get_service)):
    return svc.get_summary()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
